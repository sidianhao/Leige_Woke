from random import randint
import openpyxl,datetime


def make_demo():
    demo = openpyxl.load_workbook('心算表.xlsx')
    fution = demo.worksheets[0]
    symbol = ['+', '-', 'x', '÷']
    for row in range(1, 18, 2):
        for column in range(1, 18, 3):
            num = randint(0, 3)
            if num == 0:
                x = randint(100, 1000)
                y = randint(100, 1000)
            elif num == 1:
                x = randint(100, 1000)
                y = randint(100, 1000)
                if x < y:
                    x, y = y, x
            elif num == 2:
                x = randint(30, 300)
                y = randint(3, 9)
            else:
                x = randint(1, 100)
                y = randint(2, 9)
                while x % y != 0:
                    x = randint(1, 100)
                    y = randint(2, 9)
            mes = str(x) + symbol[num] + str(y) + '='
            fution.cell(row=row, column=column, value=mes)
            fution.cell(row=row, column=column + 1, value='')
            fution.cell(row=row, column=column + 2, value='')
            fution.cell(row=row + 1, column=column, value='')
            fution.cell(row=row + 1, column=column + 1, value='')
            fution.cell(row=row + 1, column=column + 2, value='')
    demo.save('心算表.xlsx')


while True:
    name = input('请输入你的姓名：')
    mes00 = input('请按下a键盘开始制作心算测试表：')
    if mes00 == 'a':
        make_demo()
        print('请打开心算表excel文件，准备开始做题。开始做题请按1：',end='')
        mes01 = input('')
        if mes01 == '1':
            start_time = datetime.datetime.now()
            print('请开始做题，希望你认证，仔细，确保正确率，加油！')
            print('检查、完成试卷，关闭心算表excel后，请按2键提交答案:',end='')
            mes02 = input('')
            if mes02 == '2':
                end_time = datetime.datetime.now()
                pass_time= end_time - start_time
                break
pass_time_list = str(pass_time).split(':')
# print(pass_time_list)



def check_demo():
    demo = openpyxl.load_workbook('心算表.xlsx')
    fution = demo.worksheets[0]
    true_num = 0
    false_num = 0
    false_list = []
    mess = [name+',你很棒，全对了，请继续保持！', name+',要加油哦，还差一点点就全对了。', name+',你这次表现有点糟糕，下次努力哦！']
    for row in range(1, 18, 2):
        for column in range(1, 18, 3):
            result = fution.cell(row=row, column=column + 1).value
            true_str = fution.cell(row=row, column=column).value
            true = true_str.strip('=')
            true = true.replace('x', '*')
            true = true.replace('÷', '/')
            true_result = eval(true)
            if result == true_result:
                true_num += 1
            elif result == None:
                false_num += 1
                result = '未作答'
                false_list.append(fution.cell(row=row, column=column).value + result)
            else:
                false_num += 1
                false_list.append(fution.cell(row=row, column=column).value + str(result))
    accuracy = true_num * 100 / (true_num + false_num)
    if accuracy == 100:
        mes = mess[0]
    elif accuracy > 95:
        mes = mess[1]
    else:
        mes = mess[2]
    print(name + ',你本次测试一共有{}题，耗时{}分{}秒，你回答正确{}题，错误{}题，正确率{:.1f}%。' .format
          (true_num + false_num,pass_time_list[1],pass_time_list[-1][0:2],true_num, false_num, accuracy) + mes)
    if false_num != 0:
        print('你本次测试错误的题目如下：', false_list)
    with open(r'做题记录.txt', 'a') as f:
        time = str(datetime.datetime.now())
        newtime = time.split('.')
        newtime = newtime[0]
        f.write(newtime + '(' + name + ')' + '进行了一次练习。\n' + '本次测试一共有{}题，耗时{}分{}秒，你回答正确{}题，错误{}题，正确率{:.1f}%。'.format
        (true_num + false_num, pass_time_list[1], pass_time_list[-1][0:2], true_num, false_num, accuracy) + mes + '\n\n')


check_demo()

